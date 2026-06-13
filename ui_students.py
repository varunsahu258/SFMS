"""Student master-data management screens for SFMS."""

from __future__ import annotations

import os
import re
import sqlite3
import tkinter as tk
from datetime import datetime, timedelta
from tkinter import filedialog, messagebox, simpledialog, ttk

from ui_theme import apply_theme
from ui_date import DateEntry

import auth
from ui_workspace import WorkspacePage
from config import SPLASH_BG, SPLASH_FG, STATUS_ACTIVE
from ledger import active_academic_year, charge_rows
from ledger_service import LedgerService
from security_utils import display_aadhaar
from ui_master_utils import audit, connect_db, ensure_permission_write
from utils import format_currency, now_str

CLASS_MAP = {
    "NUR": "Nursery",
    "NURSARY": "Nursery",
    "NURSERY": "Nursery",
    "KG 1": "KG-I",
    "KG IST": "KG-I",
    "KG I": "KG-I",
    "KG 2": "KG-II",
    "KG IIND": "KG-II",
    "KG II": "KG-II",
    "I": "Class 1",
    "IST": "Class 1",
    "1ST": "Class 1",
    "II": "Class 2",
    "III": "Class 3",
    "IV": "Class 4",
    "V": "Class 5",
    "VI": "Class 6",
    "VII": "Class 7",
    "VIII": "Class 8",
    "IX": "Class 9",
    "X": "Class 10",
}
PROMOTION_ORDER = [
    "Nursery", "KG-I", "KG-II", "Class 1", "Class 2", "Class 3", "Class 4",
    "Class 5", "Class 6", "Class 7", "Class 8", "Class 9", "Class 10",
]
FEE_HEADS = (
    ("Admission Fee", "BIG"),
    ("Tuition Fee", "BOTH"),
    ("Term Exam Fee", "BIG"),
    ("Computer Fee", "BIG"),
    ("Sports & Activity Fee", "BIG"),
    ("Vehicle Fee", "SMALL"),
)
FEE_SEED = {
    "Nursery": (1000, 7600, 400, 0, 300),
    "KG-I": (1000, 7800, 400, 0, 300),
    "KG-II": (1000, 7800, 400, 0, 300),
    "Class 1": (2000, 8800, 500, 300, 400),
    "Class 2": (2000, 8800, 500, 300, 400),
    "Class 3": (2000, 9000, 500, 300, 400),
    "Class 4": (2000, 9000, 500, 300, 400),
    "Class 5": (2000, 9000, 500, 300, 400),
    "Class 6": (2000, 9900, 500, 300, 500),
    "Class 7": (2000, 9900, 500, 300, 500),
    "Class 8": (2000, 9900, 500, 300, 500),
    "Class 9": (3000, 11400, 600, 300, 500),
    "Class 10": (3000, 11400, 600, 300, 500),
}
VEHICLE_FEES = {"BARELI": 3000, "KAMTONE": 3600, "PIPARIYA": 4200}
AADHAAR_RE = re.compile(r"^\d{12}$")
PHONE_RE = re.compile(r"^\d{10}$")


def student_dues_rows(conn: sqlite3.Connection, student_id: int) -> list[dict]:
    """Return authoritative active-year itemized dues for one student."""
    rows = LedgerService(conn).get_all_outstanding(active_academic_year(conn))
    return [{"fee_head": row["fee_head"], "amount_due": float(row["original_amount"]),
             "paid": float(row["paid"]), "adjustments": float(row["adjustments"]),
             "balance": float(row["outstanding"])}
            for row in rows if row["student_id"] == student_id]


def issue_student_tc(conn: sqlite3.Connection, student_id: int, override_dues=False, override_reason="") -> str:
    """Generate a TC and archive the student without deleting history."""
    from report_generator import transfer_certificate

    old_row = conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()
    if old_row is None:
        raise ValueError("Student was not found.")
    old = dict(old_row) if hasattr(old_row, "keys") else {"id": student_id}
    path = transfer_certificate(conn, student_id, override_dues, override_reason)
    conn.execute("UPDATE students SET status = 'LEFT', is_active = 0 WHERE id = ?", (student_id,))
    audit(
        conn, "STUDENT_LEFT", "students", student_id, old,
        {"status": "LEFT", "is_active": 0, "tc_issued": True},
    )
    return path


def reactivate_student(conn: sqlite3.Connection, student_id: int, reason: str) -> None:
    """Reactivate an archived student and audit the mandatory reason."""
    reason = str(reason or "").strip()
    if not reason:
        raise ValueError("A reason is mandatory.")
    row = conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()
    if row is None:
        raise ValueError("Student was not found.")
    old = dict(row) if hasattr(row, "keys") else {"id": student_id}
    conn.execute("UPDATE students SET status = 'ACTIVE', is_active = 1 WHERE id = ?", (student_id,))
    audit(conn, "STUDENT_REACTIVATE", "students", student_id, old, {"status": "ACTIVE", "is_active": 1, "reason": reason})


class StudentWindow(WorkspacePage):
    """Manage non-financial student records as an administrator or accountant."""

    @auth.require_permission("manage_students")
    def __init__(self, master=None, *, embedded: bool = False):
        """Create the student management window."""
        super().__init__(master, embedded=embedded)
        self.title("Students")
        self.geometry("1240x620")
        self.configure(bg=SPLASH_BG)
        self.search_var = tk.StringVar()
        self.filter_class_var = tk.StringVar(value="All Classes")
        self.filter_gender_var = tk.StringVar(value="All Genders")
        self.filter_rte_var = tk.StringVar(value="All RTE")
        self._ensure_import_columns()
        self._build_widgets()
        self.refresh()

    def _ensure_import_columns(self) -> None:
        """Add optional school-import columns if this database does not have them yet."""
        with connect_db() as conn:
            existing = {row["name"] for row in conn.execute("PRAGMA table_info(students)")}
            for column, ddl in {
                "scholar_no": "TEXT", "ekyc_status": "TEXT DEFAULT 'PENDING'",
                "serial_no": "TEXT", "father_name": "TEXT", "mother_name": "TEXT",
                "address": "TEXT", "dob": "TEXT", "admission_date": "TEXT",
                "mobile2": "TEXT", "sssm_id": "TEXT", "gender": "TEXT",
                "category": "TEXT", "route": "TEXT", "vehicle_fee": "REAL DEFAULT 0",
                "has_vehicle_fee": "INTEGER DEFAULT 0", "is_rte": "INTEGER NOT NULL DEFAULT 0",
                "father_education": "TEXT", "father_occupation": "TEXT",
                "family_annual_income": "REAL", "mother_education": "TEXT",
                "mother_occupation": "TEXT", "conveyance_details": "TEXT",
                "bank_account_number": "TEXT", "ifsc_code": "TEXT",
            }.items():
                if column not in existing:
                    conn.execute(f"ALTER TABLE students ADD COLUMN {column} {ddl}")

    def _build_widgets(self) -> None:
        """Build shared search, active/archive tabs, and actions."""
        top = tk.Frame(self, bg=SPLASH_BG)
        top.pack(fill="x", padx=12, pady=10)
        tk.Label(top, text="Search", bg=SPLASH_BG, fg=SPLASH_FG).pack(side="left")
        entry = ttk.Entry(top, textvariable=self.search_var, width=28)
        entry.pack(side="left", padx=8)
        entry.bind("<KeyRelease>", lambda _event: self.refresh())
        tk.Label(top, text="Class", bg=SPLASH_BG, fg=SPLASH_FG).pack(side="left", padx=(8, 2))
        self.class_filter = ttk.Combobox(top, textvariable=self.filter_class_var, state="readonly", width=14)
        self.class_filter.pack(side="left")
        self.class_filter.bind("<<ComboboxSelected>>", lambda _event: self.refresh())
        tk.Label(top, text="Gender", bg=SPLASH_BG, fg=SPLASH_FG).pack(side="left", padx=(8, 2))
        gender_filter = ttk.Combobox(top, textvariable=self.filter_gender_var, values=("All Genders", "Male", "Female", "Other"), state="readonly", width=12)
        gender_filter.pack(side="left")
        gender_filter.bind("<<ComboboxSelected>>", lambda _event: self.refresh())
        tk.Label(top, text="RTE", bg=SPLASH_BG, fg=SPLASH_FG).pack(side="left", padx=(8, 2))
        rte_filter = ttk.Combobox(top, textvariable=self.filter_rte_var, values=("All RTE", "RTE", "Non-RTE"), state="readonly", width=10)
        rte_filter.pack(side="left")
        rte_filter.bind("<<ComboboxSelected>>", lambda _event: self.refresh())
        ttk.Button(top, text="Clear", command=self._clear_search).pack(side="left", padx=6)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        active_tab = tk.Frame(self.notebook, bg=SPLASH_BG)
        archive_tab = tk.Frame(self.notebook, bg=SPLASH_BG)
        self.notebook.add(active_tab, text="Students")
        self.notebook.add(archive_tab, text="Archived Students")

        columns = ("scholar_no", "name", "class", "section", "phone", "ekyc", "status")
        self.tree = ttk.Treeview(active_tab, columns=columns, show="headings", selectmode="extended")
        for column, heading, width in (
            ("scholar_no", "Scholar No", 100), ("name", "Student Name", 220), ("class", "Class", 110),
            ("section", "Section", 80), ("phone", "Mobile 1", 110), ("ekyc", "eKYC", 90), ("status", "Status", 90),
        ):
            self.tree.heading(column, text=heading)
            self.tree.column(column, width=width)
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", lambda _event: self.edit_selected())

        archived_columns = ("name", "class", "left_on", "tc_issued", "total_paid")
        self.archived_tree = ttk.Treeview(archive_tab, columns=archived_columns, show="headings", selectmode="extended")
        for column, heading, width in (
            ("name", "Name", 250), ("class", "Class", 130),
            ("left_on", "Left On", 145), ("tc_issued", "TC Issued", 100),
            ("total_paid", "Total Paid", 130),
        ):
            self.archived_tree.heading(column, text=heading)
            self.archived_tree.column(column, width=width)
        self.archived_tree.pack(fill="both", expand=True)
        ttk.Button(archive_tab, text="Reactivate", command=self.reactivate_archived).pack(pady=8)

        buttons = tk.Frame(self, bg=SPLASH_BG)
        buttons.pack(fill="x", padx=12, pady=(0, 12))
        for text, command in (
            ("Add Student", self.add_student), ("Edit", self.edit_selected),
            ("Deactivate", self.deactivate_selected), ("Mark as Left", self.mark_left_selected),
            ("Transfer Certificate", self.transfer_certificate_selected),
            ("ID Card", self.generate_id_cards), ("Select All in Class", self.select_all_in_class),
            ("Bulk Import", self.bulk_import), ("Promote Class", self.promote_class),
        ):
            button = ttk.Button(buttons, text=text, command=command)
            button.pack(side="left", padx=3)
            if text == "Transfer Certificate":
                self.tc_button = button
        self.tree.bind("<<TreeviewSelect>>", lambda _event: self._update_tc_button())
        self._update_tc_button()

    def _update_tc_button(self) -> None:
        """Enable TC issuance only for one selected ACTIVE student."""
        selected = self.tree.selection()
        enabled = False
        if len(selected) == 1:
            values = self.tree.item(selected[0], "values")
            enabled = bool(values and values[6] == "ACTIVE")
        self.tc_button.configure(state="normal" if enabled else "disabled")

    def _clear_search(self) -> None:
        """Clear the shared active/archive search and reload both tabs."""
        auth.touch_session()
        self.search_var.set("")
        self.filter_class_var.set("All Classes")
        self.filter_gender_var.set("All Genders")
        self.filter_rte_var.set("All RTE")
        self.refresh()

    def refresh(self) -> None:
        """Reload active and archived student lists using the shared search."""
        auth.touch_session()
        for tree in (self.tree, self.archived_tree):
            for item in tree.get_children():
                tree.delete(item)
        term = f"%{self.search_var.get().strip()}%"
        filters = ["(name LIKE ? OR class LIKE ? OR aadhaar LIKE ? OR address LIKE ? OR father_name LIKE ? OR mother_name LIKE ?)"]
        params: list[object] = [term, term, term, term, term, term]
        if self.filter_class_var.get() != "All Classes":
            filters.append("class=?")
            params.append(self.filter_class_var.get())
        if self.filter_gender_var.get() != "All Genders":
            filters.append("gender=?")
            params.append(self.filter_gender_var.get())
        if self.filter_rte_var.get() != "All RTE":
            filters.append("COALESCE(is_rte,0)=?")
            params.append(1 if self.filter_rte_var.get() == "RTE" else 0)
        filter_sql = " AND ".join(filters)
        archived_filter_sql = filter_sql
        for source, target in (
            ("class", "s.class"), ("aadhaar", "s.aadhaar"), ("address", "s.address"),
            ("father_name", "s.father_name"), ("mother_name", "s.mother_name"),
            ("gender", "s.gender"), ("is_rte", "s.is_rte"),
        ):
            archived_filter_sql = archived_filter_sql.replace(source, target)
        archived_filter_sql = archived_filter_sql.replace("(name LIKE", "(s.name LIKE")
        with connect_db() as conn:
            classes = ["All Classes", *[row[0] for row in conn.execute("SELECT name FROM classes WHERE is_active=1 ORDER BY name")]]
            if hasattr(self, "class_filter"):
                self.class_filter.configure(values=classes)
            active_rows = conn.execute(
                f"""
                SELECT id, scholar_no, name, class, section, phone, ekyc_status,
                       CASE WHEN is_active = 1 THEN status ELSE 'INACTIVE' END AS status
                FROM students
                WHERE status <> 'LEFT' AND {filter_sql}
                ORDER BY class, name
                """,
                params,
            ).fetchall()
            archived_rows = conn.execute(
                f"""
                SELECT s.id, s.name, s.class,
                       COALESCE((
                           SELECT a.timestamp FROM audit_log a
                           WHERE a.record_id = CAST(s.id AS TEXT)
                             AND a.action IN ('STUDENT_LEFT', 'TC_ISSUED')
                           ORDER BY a.id DESC LIMIT 1
                       ), '') AS left_on,
                       CASE WHEN EXISTS (
                           SELECT 1 FROM audit_log a
                           WHERE a.record_id = CAST(s.id AS TEXT) AND a.action = 'TC_ISSUED'
                       ) THEN 'Yes' ELSE 'No' END AS tc_issued,
                       COALESCE((SELECT SUM(CASE WHEN p.note LIKE 'VOID of %' THEN p.amount_paid WHEN UPPER(p.payment_mode)<>'CHEQUE' OR p.cheque_status='CLEARED' THEN p.amount_paid ELSE 0 END) FROM payments p WHERE p.student_id=s.id),0) AS total_paid
                FROM students s
                WHERE s.status = 'LEFT' AND {archived_filter_sql}
                ORDER BY s.class, s.name
                """,
                params,
            ).fetchall()
        for row in active_rows:
            self.tree.insert("", "end", iid=str(row["id"]), values=(
                row["scholar_no"], row["name"], row["class"], row["section"],
                row["phone"], row["ekyc_status"], row["status"],
            ))
        for row in archived_rows:
            self.archived_tree.insert(
                "", "end", iid=str(row["id"]),
                values=(row["name"], row["class"], row["left_on"], row["tc_issued"], format_currency(row["total_paid"] or 0)),
            )

    def _selected_id(self) -> int | None:
        """Return the selected student id, if any."""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Select student", "Please select a student first.")
            return None
        return int(selected[0])

    @auth.require_permission("manage_students")
    def add_student(self) -> None:
        """Open the add-student dialog."""
        AddStudentDialog(self, on_saved=self.refresh)

    @auth.require_permission("manage_students")
    def edit_selected(self) -> None:
        """Open the edit dialog for the selected student."""
        student_id = self._selected_id()
        if student_id is not None:
            EditStudentDialog(self, student_id, on_saved=self.refresh)

    @auth.require_permission("manage_students")
    def deactivate_selected(self) -> None:
        """Deactivate the selected student after warning about unpaid dues."""
        student_id = self._selected_id()
        if student_id is None or not ensure_permission_write("manage_students"):
            return
        with connect_db() as conn:
            due = LedgerService(conn).get_outstanding(student_id, academic_year_id=None)
            if due and not messagebox.askyesno("Unpaid dues", f"Student has unpaid dues of Rs. {due:,.2f}. Deactivate anyway?"):
                return
            old = dict(conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone())
            conn.execute("UPDATE students SET is_active = 0 WHERE id = ?", (student_id,))
            audit(conn, "STUDENT_DEACTIVATE", "students", student_id, old, {"is_active": 0})
        self.refresh()

    @auth.require_permission("manage_students")
    def mark_left_selected(self) -> None:
        """Mark the selected student as LEFT if no balance remains."""
        student_id = self._selected_id()
        if student_id is None or not ensure_permission_write("manage_students"):
            return
        with connect_db() as conn:
            due = LedgerService(conn).get_outstanding(student_id, academic_year_id=None)
            if due > 0:
                messagebox.showerror("Cannot mark left", f"Student has unpaid dues of Rs. {due:,.2f}.")
                return
            old = dict(conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone())
            conn.execute("UPDATE students SET status = 'LEFT', is_active = 0 WHERE id = ?", (student_id,))
            audit(conn, "STUDENT_LEFT", "students", student_id, old, {"status": "LEFT", "is_active": 0})
        self.refresh()

    def _selected_ids(self) -> list[int]:
        """Return all selected active student IDs."""
        return [int(item) for item in self.tree.selection()]

    @auth.require_permission("manage_students")
    def transfer_certificate_selected(self) -> None:
        """Issue a TC for one active student, requesting override when required."""
        selected = self._selected_ids()
        if len(selected) != 1:
            messagebox.showwarning("Transfer Certificate", "Select exactly one active student.", parent=self)
            return
        student_id = selected[0]
        with connect_db() as conn:
            student = conn.execute("SELECT is_active, status FROM students WHERE id = ?", (student_id,)).fetchone()
            if student is None or not student["is_active"] or student["status"] != "ACTIVE":
                messagebox.showerror("Transfer Certificate", "TC is enabled only for ACTIVE students.", parent=self)
                return
            total_dues = float(LedgerService(conn).get_outstanding(student_id, academic_year_id=None) or 0)
            if total_dues > 0:
                if not auth.can_override_financial_data():
                    messagebox.showerror(
                        "Transfer Certificate",
                        "This student has unpaid dues. Only an administrator can override dues clearance.",
                        parent=self,
                    )
                    return
                DuesClearanceDialog(self, student_id, on_issued=self._tc_issued)
                return
            try:
                path = issue_student_tc(conn, student_id)
            except Exception as exc:
                messagebox.showerror("Transfer Certificate", str(exc), parent=self)
                return
        self._tc_issued(path)

    def _tc_issued(self, path: str) -> None:
        """Refresh the archive and open a generated transfer certificate."""
        self.refresh()
        if hasattr(os, "startfile"):
            os.startfile(path)
        messagebox.showinfo("Transfer Certificate", f"TC saved to:\n{path}", parent=self)

    def select_all_in_class(self) -> None:
        """Select every visible active student in the focused student's class."""
        focused = self.tree.focus() or (self.tree.selection()[0] if self.tree.selection() else "")
        if not focused:
            messagebox.showwarning("ID Cards", "Select one student to identify the class.", parent=self)
            return
        class_name = self.tree.item(focused, "values")[2]
        matches = [item for item in self.tree.get_children() if self.tree.item(item, "values")[2] == class_name]
        self.tree.selection_set(matches)

    def generate_id_cards(self) -> None:
        """Generate privacy-safe ID cards for all selected active students."""
        selected = self._selected_ids()
        if not selected:
            messagebox.showwarning("ID Cards", "Select one or more students.", parent=self)
            return
        if not messagebox.askyesno("ID Cards", f"Generate ID cards for {len(selected)} students?", parent=self):
            return
        try:
            from report_generator import student_id_card
            with connect_db() as conn:
                path = student_id_card(conn, selected)
        except Exception as exc:
            messagebox.showerror("ID Cards", str(exc), parent=self)
            return
        if hasattr(os, "startfile"):
            os.startfile(path)
        messagebox.showinfo("ID Cards", f"ID cards saved to:\n{path}", parent=self)

    @auth.require_permission("manage_students")
    def reactivate_archived(self) -> None:
        """Reactivate one archived student with a mandatory audited reason."""
        selected = self.archived_tree.selection()
        if len(selected) != 1:
            messagebox.showwarning("Reactivate", "Select exactly one archived student.", parent=self)
            return
        reason = simpledialog.askstring("Reactivate Student", "Reason for reactivation:", parent=self)
        if reason is None:
            return
        reason = reason.strip()
        if not reason:
            messagebox.showerror("Reactivate", "A reason is mandatory.", parent=self)
            return
        student_id = int(selected[0])
        with connect_db() as conn:
            reactivate_student(conn, student_id, reason)
        self.refresh()

    @auth.require_permission("manage_students")
    def bulk_import(self) -> None:
        """Open the bulk-import dialog."""
        BulkImportDialog(self, on_imported=self.refresh)

    @auth.require_permission("manage_students")
    def promote_class(self) -> None:
        """Open the class-promotion dialog."""
        PromoteClassDialog(self, on_saved=self.refresh)


class DuesClearanceDialog(tk.Toplevel):
    """Show itemized dues and permit an administrator TC override."""

    def __init__(self, master, student_id: int, on_issued=None):
        super().__init__(master)
        apply_theme(self)
        self.student_id = student_id
        self.on_issued = on_issued
        self.reason_var = tk.StringVar()
        self.title("Dues Clearance for Transfer Certificate")
        self.geometry("720x430")
        self.transient(master)
        self.grab_set()
        columns = ("fee_head", "amount_due", "paid", "balance")
        tree = ttk.Treeview(self, columns=columns, show="headings", height=11)
        for column, heading, width in (("fee_head", "Fee Head", 220), ("amount_due", "Amount Due", 140), ("paid", "Paid", 130), ("balance", "Balance", 140)):
            tree.heading(column, text=heading)
            tree.column(column, width=width)
        with connect_db() as conn:
            rows = student_dues_rows(conn, student_id)
        for row in rows:
            tree.insert("", "end", values=(row["fee_head"], format_currency(row["amount_due"]), format_currency(row["paid"]), format_currency(row["balance"])))
        tree.pack(fill="both", expand=True, padx=12, pady=12)
        reason = tk.Frame(self)
        reason.pack(fill="x", padx=12)
        tk.Label(reason, text="Override Reason").pack(side="left")
        ttk.Entry(reason, textvariable=self.reason_var).pack(side="left", padx=8, fill="x", expand=True)
        buttons = tk.Frame(self)
        buttons.pack(pady=12)
        ttk.Button(buttons, text="Override and Issue TC", command=self.issue).pack(side="left", padx=5)
        ttk.Button(buttons, text="Cancel", command=self.destroy).pack(side="left", padx=5)

    @auth.require_role("ADMIN")
    def issue(self) -> None:
        reason = self.reason_var.get().strip()
        if not reason:
            messagebox.showerror("Dues Override", "A reason is mandatory.", parent=self)
            return
        try:
            with connect_db() as conn:
                path = issue_student_tc(conn, self.student_id, True, reason)
        except Exception as exc:
            messagebox.showerror("Transfer Certificate", str(exc), parent=self)
            return
        self.destroy()
        if self.on_issued:
            self.on_issued(path)


class StudentDialog(tk.Toplevel):
    """Base dialog for the complete student admission profile."""

    FIELD_ROWS = (
        ("Scholar No", "scholar_no"), ("eKYC Status", "ekyc_status"),
        ("SL. No.", "serial_no"), ("Student Name", "name"),
        ("Father's Name", "father_name"), ("Mother's Name", "mother_name"),
        ("Address", "address"), ("Date of Birth", "dob"),
        ("Admission Date", "admission_date"), ("Class", "class"),
        ("Section", "section"), ("Mobile No. 1", "phone"),
        ("Mobile No. 2", "mobile2"), ("SSSM ID", "sssm_id"),
        ("Gender", "gender"), ("Aadhaar Card No.", "aadhaar"),
        ("Category", "category"), ("RTE Student", "is_rte"),
        ("Father's Education", "father_education"), ("Father's Occupation", "father_occupation"),
        ("Family Annual Income", "family_annual_income"),
        ("Mother's Education", "mother_education"), ("Mother's Occupation", "mother_occupation"),
        ("Conveyance Details", "conveyance_details"),
        ("Bank Account Number", "bank_account_number"), ("IFSC Code", "ifsc_code"),
    )

    def __init__(self, master, title: str, on_saved=None):
        super().__init__(master)
        apply_theme(self)
        self.on_saved = on_saved
        self.title(title)
        self.geometry("840x760")
        self.configure(bg=SPLASH_BG)
        self.vars = {key: tk.StringVar() for _label, key in self.FIELD_ROWS}
        self.vars["ekyc_status"].set("PENDING")
        self._build_form()

    def _classes(self) -> list[str]:
        with connect_db() as conn:
            return [row[0] for row in conn.execute("SELECT name FROM classes WHERE is_active=1 ORDER BY name")]

    def _sections(self, class_name: str | None = None) -> list[str]:
        class_name = class_name if class_name is not None else self.vars["class"].get()
        with connect_db() as conn:
            return [row[0] for row in conn.execute(
                """SELECT s.name FROM sections s JOIN classes c ON c.id=s.class_id
                   WHERE c.name=? AND c.is_active=1 AND s.is_active=1 ORDER BY s.name""",
                (class_name,),
            )]

    def _class_changed(self, _event=None) -> None:
        values = self._sections()
        self.section_combo.configure(values=values)
        if self.vars["section"].get() not in values:
            self.vars["section"].set("")

    def _build_form(self) -> None:
        outer = tk.Frame(self, bg=SPLASH_BG)
        outer.pack(fill="both", expand=True, padx=18, pady=14)
        canvas = tk.Canvas(outer, bg=SPLASH_BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        frame = tk.Frame(canvas, bg=SPLASH_BG)
        frame.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        for index, (label, key) in enumerate(self.FIELD_ROWS):
            row, pair = divmod(index, 2)
            column = pair * 2
            tk.Label(frame, text=label, bg=SPLASH_BG, fg=SPLASH_FG).grid(row=row, column=column, sticky="w", padx=(4, 8), pady=6)
            if key == "class":
                widget = ttk.Combobox(frame, textvariable=self.vars[key], values=self._classes(), state="readonly", width=24)
                widget.bind("<<ComboboxSelected>>", self._class_changed)
            elif key == "section":
                widget = ttk.Combobox(frame, textvariable=self.vars[key], state="readonly", width=24)
                self.section_combo = widget
            elif key == "ekyc_status":
                widget = ttk.Combobox(frame, textvariable=self.vars[key], values=("PENDING", "VERIFIED", "FAILED", "NOT REQUIRED"), state="readonly", width=24)
            elif key == "gender":
                widget = ttk.Combobox(frame, textvariable=self.vars[key], values=("Male", "Female", "Other"), state="readonly", width=24)
            elif key == "category":
                widget = ttk.Combobox(frame, textvariable=self.vars[key], values=("OBC", "SC", "ST", "General"), state="readonly", width=24)
            elif key == "is_rte":
                widget = ttk.Combobox(frame, textvariable=self.vars[key], values=("No", "Yes"), state="readonly", width=24)
                if not self.vars[key].get():
                    self.vars[key].set("No")
            elif key in {"dob", "admission_date"}:
                widget = DateEntry(frame, textvariable=self.vars[key], width=21)
            else:
                widget = ttk.Entry(frame, textvariable=self.vars[key], width=27)
            widget.grid(row=row, column=column + 1, sticky="ew", padx=(0, 16), pady=6)
            if key == "aadhaar":
                self.aadhaar_entry = widget
        for column in (1, 3):
            frame.columnconfigure(column, weight=1)
        ttk.Button(frame, text="Save", command=self.save).grid(
            row=(len(self.FIELD_ROWS) + 1) // 2, column=0, columnspan=4, pady=18
        )

    def _validate(self, conn: sqlite3.Connection, student_id: int | None = None) -> bool:
        required = (("scholar_no", "Scholar number"), ("name", "Student name"), ("class", "Class"), ("section", "Section"))
        for key, label in required:
            if not self.vars[key].get().strip():
                messagebox.showerror("Validation", f"{label} is required.", parent=self)
                return False
        aadhaar = re.sub(r"\D", "", self.vars["aadhaar"].get())
        if aadhaar and not AADHAAR_RE.match(aadhaar):
            messagebox.showerror("Validation", "Aadhaar must be exactly 12 digits when provided.", parent=self)
            return False
        for key, label in (("phone", "Mobile No. 1"), ("mobile2", "Mobile No. 2")):
            phone = re.sub(r"\D", "", self.vars[key].get())
            if phone and not PHONE_RE.match(phone):
                messagebox.showerror("Validation", f"{label} must be exactly 10 digits when provided.", parent=self)
                return False
            self.vars[key].set(phone)
        for column, value, label in (("scholar_no", self.vars["scholar_no"].get().strip(), "Scholar number"), ("aadhaar", aadhaar, "Aadhaar")):
            if not value:
                continue
            sql = f"SELECT id FROM students WHERE {column}=?"
            params: list[object] = [value]
            if student_id is not None:
                sql += " AND id<>?"
                params.append(student_id)
            if conn.execute(sql, params).fetchone():
                messagebox.showerror("Validation", f"{label} already exists.", parent=self)
                return False
        self.vars["aadhaar"].set(aadhaar)
        income = self.vars.get("family_annual_income")
        if income is not None and income.get().strip():
            try:
                if float(income.get().strip()) < 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Validation", "Family annual income must be a positive number when provided.", parent=self)
                return False
        ifsc = self.vars.get("ifsc_code")
        if ifsc is not None:
            ifsc.set(ifsc.get().strip().upper())
            if ifsc.get() and not re.match(r"^[A-Z]{4}0[A-Z0-9]{6}$", ifsc.get()):
                messagebox.showerror("Validation", "IFSC code is invalid.", parent=self)
                return False
        return True

    def values(self) -> dict[str, str | None]:
        values = {key: var.get().strip() for key, var in self.vars.items()}
        values["is_rte"] = 1 if str(values.get("is_rte", "")).strip().lower() in {"1", "yes", "true", "rte"} else 0
        values["aadhaar"] = values["aadhaar"] or None
        values["phone"] = values["phone"] or None
        values["mobile2"] = values["mobile2"] or None
        values["family_annual_income"] = values["family_annual_income"] or None
        return values

    def save(self) -> None:
        raise NotImplementedError


class AddStudentDialog(StudentDialog):
    def __init__(self, master, on_saved=None):
        super().__init__(master, "Add Student", on_saved)

    @auth.require_permission("manage_students")
    def save(self) -> None:
        if not ensure_permission_write("manage_students"):
            return
        with connect_db() as conn:
            if not self._validate(conn):
                return
            values = self.values()
            columns = tuple(values)
            cursor = conn.execute(
                f"INSERT INTO students ({','.join(columns)},guardian_name,is_active,status,created_at) "
                f"VALUES ({','.join('?' for _ in columns)},?,1,?,?)",
                (*values.values(), values["father_name"] or "", STATUS_ACTIVE, now_str()),
            )
            audit(conn, "STUDENT_ADD", "students", cursor.lastrowid, None, values)
        if self.on_saved:
            self.on_saved()
        self.destroy()


class EditStudentDialog(StudentDialog):
    def __init__(self, master, student_id: int, on_saved=None):
        self.student_id = student_id
        super().__init__(master, "Edit Student", on_saved)
        self._load()

    def _load(self) -> None:
        with connect_db() as conn:
            row = conn.execute("SELECT * FROM students WHERE id=?", (self.student_id,)).fetchone()
        if row:
            for key in self.vars:
                if key == "is_rte":
                    self.vars[key].set("Yes" if row[key] else "No")
                else:
                    self.vars[key].set(row[key] or "")
            self.section_combo.configure(values=self._sections(row["class"] or ""))

    @auth.require_permission("manage_students")
    def save(self) -> None:
        if not ensure_permission_write("manage_students"):
            return
        with connect_db() as conn:
            if not self._validate(conn, self.student_id):
                return
            old = dict(conn.execute("SELECT * FROM students WHERE id=?", (self.student_id,)).fetchone())
            values = self.values()
            assignments = ",".join(f"{column}=?" for column in values)
            conn.execute(
                f"UPDATE students SET {assignments},guardian_name=? WHERE id=?",
                (*values.values(), values["father_name"] or "", self.student_id),
            )
            audit(conn, "STUDENT_EDIT", "students", self.student_id, old, values)
        if self.on_saved:
            self.on_saved()
        self.destroy()


class BulkImportDialog(tk.Toplevel):
    """Import students using the same columns as the student details form."""

    HEADERS = (
        "SCHOLAR NO", "EKYC STATUS", "SL. NO.", "STUDENTS NAME", "FATHERS NAME",
        "MOTHERS NAME", "ADDRESS", "DOB", "ADMISION DATE", "CLASS", "SECTION",
        "MOBILE NO 1", "MOBILE NO 2", "SSSM ID", "GENDER", "AADHAR CARD NO", "CATEGORY",
    )
    FIELD_MAP = {
        "SCHOLAR NO": "scholar_no", "EKYC STATUS": "ekyc_status", "SL. NO.": "serial_no",
        "STUDENTS NAME": "name", "FATHERS NAME": "father_name", "MOTHERS NAME": "mother_name",
        "ADDRESS": "address", "DOB": "dob", "ADMISION DATE": "admission_date",
        "ADMISSION DATE": "admission_date", "CLASS": "class", "SECTION": "section",
        "MOBILE NO 1": "phone", "MOBILE NO 2": "mobile2", "SSSM ID": "sssm_id",
        "GENDER": "gender", "AADHAR CARD NO": "aadhaar", "AADHAAR CARD NO": "aadhaar",
        "CATEGORY": "category",
    }

    def __init__(self, master, on_imported=None):
        super().__init__(master)
        apply_theme(self)
        self.on_imported = on_imported
        self.rows: list[dict] = []
        self.title("Bulk Import Students")
        self.geometry("1180x600")
        self.configure(bg=SPLASH_BG)
        self.file_var = tk.StringVar()
        self.summary_var = tk.StringVar(value="Download the template or select an .xlsx file.")
        self._build_widgets()

    def _build_widgets(self) -> None:
        top = tk.Frame(self, bg=SPLASH_BG)
        top.pack(fill="x", padx=12, pady=10)
        ttk.Entry(top, textvariable=self.file_var, width=78).pack(side="left", padx=(0, 8))
        ttk.Button(top, text="Browse", command=self.browse).pack(side="left")
        ttk.Button(top, text="Preview", command=self.preview).pack(side="left", padx=6)
        ttk.Button(top, text="Download Template", command=self.download_template).pack(side="left")
        columns = ("scholar", "name", "class", "section", "father", "phone", "status")
        self.tree = ttk.Treeview(self, columns=columns, show="headings")
        for column, heading, width in (
            ("scholar", "Scholar No", 110), ("name", "Student Name", 210),
            ("class", "Class", 100), ("section", "Section", 80),
            ("father", "Father's Name", 180), ("phone", "Mobile 1", 110),
            ("status", "Validation", 300),
        ):
            self.tree.heading(column, text=heading)
            self.tree.column(column, width=width)
        self.tree.tag_configure("ok", foreground="green")
        self.tree.tag_configure("error", foreground="red")
        self.tree.pack(fill="both", expand=True, padx=12, pady=8)
        bottom = tk.Frame(self, bg=SPLASH_BG)
        bottom.pack(fill="x", padx=12, pady=10)
        tk.Label(bottom, textvariable=self.summary_var, bg=SPLASH_BG, fg=SPLASH_FG).pack(side="left")
        ttk.Button(bottom, text="Import Valid Rows", command=self.import_valid_rows).pack(side="right")

    def download_template(self) -> None:
        from openpyxl import Workbook

        path = filedialog.asksaveasfilename(
            parent=self, defaultextension=".xlsx", initialfile="SFMS_Student_Import_Template.xlsx",
            filetypes=(("Excel workbooks", "*.xlsx"),),
        )
        if not path:
            return
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Students"
        sheet.append(self.HEADERS)
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = cell.font.copy(bold=True)
        workbook.save(path)
        messagebox.showinfo("Student Import", f"Template saved to:\n{path}", parent=self)

    def browse(self) -> None:
        path = filedialog.askopenfilename(parent=self, filetypes=(("Excel workbooks", "*.xlsx"),))
        if path:
            self.file_var.set(path)

    @staticmethod
    def _cell_text(value) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%d-%m-%Y")
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def preview(self) -> None:
        from openpyxl import load_workbook

        path = self.file_var.get().strip()
        if not path:
            messagebox.showwarning("Student Import", "Select an .xlsx file.", parent=self)
            return
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["Students"] if "Students" in workbook.sheetnames else workbook.active
        iterator = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(iterator)
        except StopIteration:
            messagebox.showerror("Student Import", "The workbook is empty.", parent=self)
            return
        headers = [self._cell_text(value).upper() for value in raw_headers]
        mapped = {index: self.FIELD_MAP[name] for index, name in enumerate(headers) if name in self.FIELD_MAP}
        missing = [name for name in self.HEADERS if self.FIELD_MAP[name] not in mapped.values()]
        if missing:
            messagebox.showerror("Student Import", "Missing columns: " + ", ".join(missing), parent=self)
            return
        with connect_db() as conn:
            classes = {row[0] for row in conn.execute("SELECT name FROM classes WHERE is_active=1")}
            sections = {(row[0], row[1]) for row in conn.execute(
                "SELECT c.name,s.name FROM sections s JOIN classes c ON c.id=s.class_id WHERE c.is_active=1 AND s.is_active=1"
            )}
            existing_scholars = {str(row[0]) for row in conn.execute("SELECT scholar_no FROM students WHERE COALESCE(scholar_no,'')<>''")}
            existing_aadhaar = {str(row[0]) for row in conn.execute("SELECT aadhaar FROM students WHERE COALESCE(aadhaar,'')<>''")}
        seen_scholars: set[str] = set()
        seen_aadhaar: set[str] = set()
        self.rows = []
        for excel_row, raw in enumerate(iterator, start=2):
            values = {field: self._cell_text(raw[index] if index < len(raw) else "") for index, field in mapped.items()}
            if not any(values.values()):
                continue
            values["aadhaar"] = re.sub(r"\D", "", values["aadhaar"])
            values["phone"] = re.sub(r"\D", "", values["phone"])
            values["mobile2"] = re.sub(r"\D", "", values["mobile2"])
            errors = []
            for key, label in (("scholar_no", "Scholar No"), ("name", "Student Name"), ("class", "Class"), ("section", "Section")):
                if not values[key]:
                    errors.append(f"{label} required")
            if values["class"] and values["class"] not in classes:
                errors.append("Class not in master")
            if values["class"] and values["section"] and (values["class"], values["section"]) not in sections:
                errors.append("Section not in class master")
            if values["scholar_no"] in existing_scholars or values["scholar_no"] in seen_scholars:
                errors.append("Duplicate Scholar No")
            if values["aadhaar"] and (not AADHAAR_RE.match(values["aadhaar"]) or values["aadhaar"] in existing_aadhaar or values["aadhaar"] in seen_aadhaar):
                errors.append("Invalid/duplicate Aadhaar")
            for key, label in (("phone", "Mobile 1"), ("mobile2", "Mobile 2")):
                if values[key] and not PHONE_RE.match(values[key]):
                    errors.append(f"Invalid {label}")
            if values["category"].upper() == "GENEREL":
                values["category"] = "General"
            values["status"] = "; ".join(errors) if errors else "OK"
            values["excel_row"] = excel_row
            self.rows.append(values)
            seen_scholars.add(values["scholar_no"])
            if values["aadhaar"]:
                seen_aadhaar.add(values["aadhaar"])
        self._render_preview()

    def _render_preview(self) -> None:
        self.tree.delete(*self.tree.get_children())
        valid = 0
        for row in self.rows:
            ok = row["status"] == "OK"
            valid += int(ok)
            self.tree.insert("", "end", values=(row["scholar_no"], row["name"], row["class"], row["section"], row["father_name"], row["phone"], row["status"]), tags=("ok" if ok else "error",))
        self.summary_var.set(f"{valid} valid, {len(self.rows) - valid} errors")

    @auth.require_permission("manage_students")
    def import_valid_rows(self) -> None:
        if not ensure_permission_write("manage_students"):
            return
        valid_rows = [row for row in self.rows if row["status"] == "OK"]
        if not valid_rows:
            messagebox.showwarning("Student Import", "There are no valid rows to import.", parent=self)
            return
        columns = tuple(self.FIELD_MAP[name] for name in self.HEADERS)
        with connect_db() as conn:
            for row in valid_rows:
                cursor = conn.execute(
                    f"INSERT INTO students ({','.join(columns)},guardian_name,is_active,status,created_at) "
                    f"VALUES ({','.join('?' for _ in columns)},?,1,?,?)",
                    (*(row[column] or None for column in columns), row["father_name"], STATUS_ACTIVE, now_str()),
                )
                audit(conn, "STUDENT_IMPORT", "students", cursor.lastrowid, None, {column: row[column] for column in columns})
        messagebox.showinfo("Student Import", f"Imported {len(valid_rows)} students.", parent=self)
        if self.on_imported:
            self.on_imported()
        self.destroy()


class PromoteClassDialog(tk.Toplevel):
    """Dialog for promoting selected students from one class to another."""

    def __init__(self, master, on_saved=None):
        """Create the promotion dialog."""
        super().__init__(master)
        apply_theme(self)
        self.on_saved = on_saved
        self.title("Promote Class")
        self.geometry("520x500")
        self.source_var = tk.StringVar()
        self.target_var = tk.StringVar()
        self._build_widgets()

    def _classes(self) -> list[str]:
        """Return classes that currently have students."""
        with connect_db() as conn:
            return [row[0] for row in conn.execute("SELECT DISTINCT class FROM students WHERE class IS NOT NULL AND class <> '' ORDER BY class")]

    def _build_widgets(self) -> None:
        """Build class selectors and student checklist."""
        top = tk.Frame(self)
        top.pack(fill="x", padx=12, pady=10)
        classes = self._classes()
        ttk.Combobox(top, textvariable=self.source_var, values=classes, state="readonly").pack(side="left", padx=4)
        ttk.Combobox(top, textvariable=self.target_var, values=classes, state="readonly").pack(side="left", padx=4)
        ttk.Button(top, text="Load", command=self.load_students).pack(side="left", padx=4)
        self.tree = ttk.Treeview(self, columns=("selected", "id", "name"), show="headings")
        for column in ("selected", "id", "name"):
            self.tree.heading(column, text=column.title())
        self.tree.pack(fill="both", expand=True, padx=12, pady=8)
        self.tree.bind("<Double-1>", self.toggle_selected)
        ttk.Button(self, text="Confirm Promotion", command=self.confirm).pack(pady=10)

    def load_students(self) -> None:
        """Load students from the selected source class."""
        auth.touch_session()
        for item in self.tree.get_children():
            self.tree.delete(item)
        with connect_db() as conn:
            rows = conn.execute("SELECT id, name FROM students WHERE class = ? AND is_active = 1 ORDER BY name", (self.source_var.get(),)).fetchall()
        for row in rows:
            self.tree.insert("", "end", iid=str(row["id"]), values=("Yes", row["id"], row["name"]))

    def toggle_selected(self, _event) -> None:
        """Toggle whether the highlighted student will be promoted."""
        auth.touch_session()
        item = self.tree.focus()
        if item:
            values = list(self.tree.item(item, "values"))
            values[0] = "No" if values[0] == "Yes" else "Yes"
            self.tree.item(item, values=values)

    @auth.require_permission("manage_students")
    def confirm(self) -> None:
        """Promote selected students to the target class and audit each update."""
        if not ensure_permission_write("manage_students"):
            return
        source = self.source_var.get()
        target = self.target_var.get()
        if not source or not target:
            messagebox.showerror("Promotion", "Select source and target classes.")
            return
        selected_ids = [int(item) for item in self.tree.get_children() if self.tree.item(item, "values")[0] == "Yes"]
        with connect_db() as conn:
            for student_id in selected_ids:
                old = dict(conn.execute("SELECT id, name, class FROM students WHERE id = ?", (student_id,)).fetchone())
                conn.execute("UPDATE students SET class = ? WHERE id = ?", (target, student_id))
                audit(conn, "CLASS_PROMOTION", "students", student_id, old, {"class": target})
        if self.on_saved:
            self.on_saved()
        messagebox.showinfo("Promotion", f"Promoted {len(selected_ids)} students from {source} to {target}.")
        self.destroy()


def normalize_class(value) -> str:
    """Normalize school-specific class names to SFMS class labels."""
    text = re.sub(r"\s+", " ", str(value).upper().replace(".", " ")).strip()
    return CLASS_MAP.get(text, text.title())


def extract_phone(value) -> str:
    """Extract the first 10-digit phone number from a cell value."""
    digits = re.sub(r"\D", "", str(value or ""))
    match = re.search(r"\d{10}", digits)
    return match.group(0) if match else ""


def normalize_date(value) -> str:
    """Normalize Excel date values and common date strings to DD-MM-YYYY text."""
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=int(value))).strftime("%d-%m-%Y")
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d-%m-%Y")
        except ValueError:
            continue
    return text


def normalize_gender(value) -> str:
    """Normalize Girl/Boy values to F/M."""
    text = str(value or "").strip().upper()
    if text == "BOY":
        return "M"
    if text == "GIRL":
        return "F"
    return ""


def normalize_category(value) -> str:
    """Normalize caste/category values to SC/ST/OBC/GEN when recognized."""
    text = str(value or "").strip().upper()
    return text if text in {"SC", "ST", "OBC", "GEN"} else text


def route_from_address(address) -> tuple[str, int]:
    """Return route code and annual vehicle fee from an address cell."""
    text = str(address or "").upper()
    if "KAMTON" in text or "KAMTONE" in text:
        return "KAMTONE", 3600
    if "PIPARIYA" in text or "SALAIYA" in text or "CHEENDMOD" in text:
        return "PIPARIYA", 4200
    return "BARELI", 3000


def header_value(values: list, headers: dict[str, int], *candidates: str):
    """Return the value for the first matching header candidate."""
    for candidate in candidates:
        candidate_lower = candidate.lower()
        for header, index in headers.items():
            if candidate_lower == header or candidate_lower in header:
                return values[index] if index < len(values) else None
    return None


def build_import_row(values: list, headers: dict[str, int], class_name: str, existing_aadhaar: set, aadhaar_seen: set) -> dict | None:
    """Build and validate a single bulk-import preview row."""
    name = str(header_value(values, headers, "Student Name") or "").strip()
    if not name:
        return None
    aadhaar = re.sub(r"\s+", "", str(header_value(values, headers, "AADHAAR CARD NO") or ""))
    phone = extract_phone(header_value(values, headers, "Mob.No.1", "Mob"))
    address = header_value(values, headers, "Address")
    route, vehicle_fee = route_from_address(address)
    has_vehicle_fee = bool(str(header_value(values, headers, "conveyance") or "").strip())
    errors = []
    if not AADHAAR_RE.match(aadhaar):
        errors.append("invalid Aadhaar")
    elif aadhaar in existing_aadhaar or aadhaar in aadhaar_seen:
        errors.append("duplicate Aadhaar")
    if not PHONE_RE.match(phone):
        errors.append("invalid phone")
    status = "OK" if not errors else "ERROR: " + ", ".join(errors)
    return {
        "name": name,
        "class": class_name,
        "section": "",
        "dob": normalize_date(header_value(values, headers, "D.O.B", "DOB")),
        "phone": phone,
        "aadhaar": aadhaar,
        "guardian_name": str(header_value(values, headers, "Father's Name", "Father") or "").strip(),
        "gender": normalize_gender(header_value(values, headers, "Girl/Boy")),
        "category": normalize_category(header_value(values, headers, "Category")),
        "route": route,
        "vehicle_fee": vehicle_fee if has_vehicle_fee else 0,
        "has_vehicle_fee": has_vehicle_fee,
        "status": status,
    }
