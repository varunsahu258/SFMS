"""Styled Excel export helpers for SFMS reports."""

from __future__ import annotations

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:  # optional Excel dependency in lightweight installs
    Workbook = Font = PatternFill = get_column_letter = None

from config import REPORTS_DIR
from security_utils import sanitize_excel_cell

NAVY = "1A1A5E"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MONEY_FORMAT = "#,##0.00"


def _is_monetary(header: str) -> bool:
    """Return whether a report header represents monetary values."""
    normalized = str(header).lower()
    return any(token in normalized for token in ("amount", "rs", "balance", "due"))


def _display_value(value) -> str:
    """Return the text used when measuring an Excel column."""
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:,.2f}"
    return str(value)


def export_to_excel(
    data: list[dict],
    headers: list[str],
    title: str,
    filename: str,
    metadata: dict,
) -> str:
    """Create a styled Data sheet and metadata Info sheet in REPORTS_DIR."""
    Path(REPORTS_DIR).mkdir(parents=True, exist_ok=True)
    output_name = filename if str(filename).lower().endswith(".xlsx") else f"{filename}.xlsx"
    output_path = str(Path(REPORTS_DIR) / output_name)

    if Workbook is None:
        with open(output_path, "w", encoding="utf-8") as handle:
            handle.write(",".join(headers) + "\n")
            for item in data:
                handle.write(",".join(sanitize_excel_cell(str(item.get(header, ""))).replace("\n", " ") for header in headers) + "\n")
        return output_path

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Data"
    info_sheet = workbook.create_sheet("Info")

    header_fill = PatternFill(fill_type="solid", fgColor=NAVY)
    alternate_fill = PatternFill(fill_type="solid", fgColor=LIGHT_GRAY)
    header_font = Font(color=WHITE, bold=True)

    for column_index, header in enumerate(headers, start=1):
        cell = data_sheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font

    monetary_columns = {
        column_index
        for column_index, header in enumerate(headers, start=1)
        if _is_monetary(header)
    }
    widths = [len(str(header)) for header in headers]
    for row_index, item in enumerate(data, start=2):
        for column_index, header in enumerate(headers, start=1):
            value = item.get(header, "")
            if isinstance(value, str):
                value = sanitize_excel_cell(value)
            cell = data_sheet.cell(row=row_index, column=column_index, value=value)
            if isinstance(value, str):
                cell.data_type = "s"
            if row_index % 2 == 1:
                cell.fill = alternate_fill
            if column_index in monetary_columns and isinstance(value, (int, float)):
                cell.number_format = MONEY_FORMAT
                display = f"{float(value):,.2f}"
            else:
                display = _display_value(value)
            widths[column_index - 1] = max(widths[column_index - 1], len(display))

    for column_index, width in enumerate(widths, start=1):
        data_sheet.column_dimensions[get_column_letter(column_index)].width = min(max(width + 2, 10), 60)
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = data_sheet.dimensions

    info_rows = [
        ("title", title),
        ("generated_by", metadata.get("generated_by", "")),
        ("generated_at", metadata.get("generated_at", "")),
        ("filters", metadata.get("filters", "")),
    ]
    for row_index, (key, value) in enumerate(info_rows, start=1):
        key_cell = info_sheet.cell(row=row_index, column=1, value=key)
        key_cell.font = Font(bold=True)
        value_cell = info_sheet.cell(row=row_index, column=2, value=sanitize_excel_cell(str(value or "")))
        value_cell.data_type = "s"
    info_sheet.column_dimensions["A"].width = 18
    info_sheet.column_dimensions["B"].width = 80

    workbook.save(output_path)
    return output_path
