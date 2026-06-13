"""PDF and Excel exports for SFMS timetable versions."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ModuleNotFoundError:  # optional export dependency in lightweight installs
    Workbook = Alignment = Font = PatternFill = None
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
except ModuleNotFoundError:  # optional export dependency in lightweight installs
    colors = A4 = landscape = mm = PageBreak = Paragraph = SimpleDocTemplate = Spacer = Table = TableStyle = None

from config import REPORTS_DIR, SCHOOL_NAME
from timetable_db import get_schedule_config, get_teacher, get_version, list_timetable, period_times, timetable_classes

DARK = colors.HexColor("#5b3fc0") if colors else "#5b3fc0"
PASTELS = ("E8F1FF", "E5F7EC", "FFF0E3", "F4E8FF", "FFF8D9", "E5F7F6")


def _stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _path(stem: str, suffix: str) -> str:
    Path(REPORTS_DIR).mkdir(parents=True, exist_ok=True)
    return str(Path(REPORTS_DIR) / f"{stem}_{_stamp()}.{suffix}")


def _settings(conn) -> dict:
    return dict(conn.execute("SELECT key,value FROM settings"))


def _version(conn, version_id: int) -> dict:
    row = get_version(conn, version_id)
    if row is None:
        raise ValueError("Timetable version was not found.")
    return row


def _grid_data(conn, version_id: int, class_name: str) -> tuple[list[list], list[str]]:
    config = get_schedule_config(conn)
    days = [day for day in config["working_days"].split(",") if day]
    times = period_times(config)
    slots = {(row["day"], int(row["period_no"])): row for row in list_timetable(conn, version_id, class_name=class_name)}
    header = ["Period"] + days
    data = [header]
    for period_no, (start, end) in enumerate(times, 1):
        row = [f"P{period_no}\n{start}-{end}"]
        for day in days:
            slot = slots.get((day, period_no), {})
            if slot.get("is_free", 1):
                row.append("FREE")
            else:
                row.append(f"{slot.get('subject_code') or slot.get('subject_name') or ''}\n{slot.get('teacher_name') or ''}")
        data.append(row)
    return data, days


def _simple_pdf(path: str, lines: list[str]) -> str:
    text_ops = ["BT /F1 9 Tf 40 560 Td"]
    for line in lines:
        safe = str(line).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        text_ops.append(f"({safe[:120]}) Tj 0 -12 Td")
    text_ops.append("ET")
    stream = "\n".join(text_ops)
    objects = [
        "1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj",
        "2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj",
        "3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 842 595] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj",
        "4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj",
        f"5 0 obj << /Length {len(stream.encode())} >> stream\n{stream}\nendstream endobj",
    ]
    offsets, body = [], "%PDF-1.4\n"
    for obj in objects:
        offsets.append(len(body.encode())); body += obj + "\n"
    xref = len(body.encode())
    body += f"xref\n0 {len(objects)+1}\n0000000000 65535 f \n" + "".join(f"{off:010d} 00000 n \n" for off in offsets)
    body += f"trailer << /Size {len(objects)+1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF"
    Path(path).write_bytes(body.encode("latin-1", "replace"))
    return path


def _pdf_document(path: str, title: str):
    if SimpleDocTemplate is None:
        return None
    return SimpleDocTemplate(path, pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm,
                             topMargin=12 * mm, bottomMargin=12 * mm, title=title, author="SFMS")


def _title_story(conn, title: str, version: dict) -> list:
    settings = _settings(conn)
    from reportlab.lib.styles import getSampleStyleSheet
    styles = getSampleStyleSheet()
    return [
        Paragraph(settings.get("school_name") or SCHOOL_NAME, styles["Title"]),
        Paragraph(title, styles["Heading2"]),
        Paragraph(f"Version: {version['label']} | Academic year: {version['academic_year']} | Generated: {version['generated_at']}", styles["BodyText"]),
        Spacer(1, 5 * mm),
    ]


def _grid_table(data: list[list]) -> Table:
    page_width = landscape(A4)[0] - 24 * mm
    widths = [28 * mm] + [(page_width - 28 * mm) / max(1, len(data[0]) - 1)] * (len(data[0]) - 1)
    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return table


def class_timetable_pdf(conn, version_id: int, class_name: str) -> str:
    version = _version(conn, version_id)
    included = {row["name"] for row in timetable_classes(conn)}
    if class_name not in included:
        raise ValueError("This class is excluded from timetable exports.")
    data, _days = _grid_data(conn, version_id, class_name)
    path = _path(f"timetable_class_{re.sub(r'[^A-Za-z0-9_-]+', '_', class_name)}", "pdf")
    if SimpleDocTemplate is None:
        return _simple_pdf(path, [f"{SCHOOL_NAME} Class Timetable — {class_name}", *[" | ".join(map(str, row)) for row in data]])
    story = _title_story(conn, f"Class Timetable — {class_name}", version)
    story.append(_grid_table(data))
    _pdf_document(path, f"Class Timetable {class_name}").build(story)
    return path


def master_timetable_pdf(conn, version_id: int) -> str:
    version = _version(conn, version_id)
    included = {row["name"] for row in timetable_classes(conn)}
    classes = [row[0] for row in conn.execute("SELECT DISTINCT class_name FROM tt_timetable WHERE version_id=? ORDER BY class_name", (version_id,)) if row[0] in included]
    if not classes:
        raise ValueError("This version has no timetable slots.")
    path = _path("timetable_master", "pdf")
    if SimpleDocTemplate is None:
        lines = [f"{SCHOOL_NAME} Master Timetable"]
        for class_name in classes:
            lines.append(f"Class Timetable — {class_name}")
            lines.extend(" | ".join(map(str, row)) for row in _grid_data(conn, version_id, class_name)[0])
        return _simple_pdf(path, lines)
    story = []
    for index, class_name in enumerate(classes):
        if index:
            story.append(PageBreak())
        story.extend(_title_story(conn, f"Class Timetable — {class_name}", version))
        story.append(_grid_table(_grid_data(conn, version_id, class_name)[0]))
    _pdf_document(path, "Master Timetable").build(story)
    return path


def teacher_duty_pdf(conn, version_id: int, teacher_id: int) -> str:
    version = _version(conn, version_id)
    teacher = get_teacher(conn, teacher_id)
    if teacher is None:
        raise ValueError("Teacher was not found.")
    config = get_schedule_config(conn)
    days = [day for day in config["working_days"].split(",") if day]
    times = period_times(config)
    slots = {(row["day"], int(row["period_no"])): row for row in list_timetable(conn, version_id, teacher_id=teacher_id)}
    data = [["Period"] + days]
    for period_no, (start, end) in enumerate(times, 1):
        data.append([f"P{period_no}\n{start}-{end}"] + [
            f"{slots[(day, period_no)]['class_name']}\n{slots[(day, period_no)].get('subject_code') or slots[(day, period_no)].get('subject_name') or ''}"
            if (day, period_no) in slots else "FREE" for day in days
        ])
    path = _path(f"teacher_duty_{teacher_id}", "pdf")
    if SimpleDocTemplate is None:
        return _simple_pdf(path, [f"{SCHOOL_NAME} Teacher Duty — {teacher['name']}", *[" | ".join(map(str, row)) for row in data]])
    story = _title_story(conn, f"Teacher Duty — {teacher['name']}", version)
    story.append(_grid_table(data))
    _pdf_document(path, f"Teacher Duty {teacher['name']}").build(story)
    return path


def timetable_excel(conn, version_id: int) -> str:
    version = _version(conn, version_id)
    included = {row["name"] for row in timetable_classes(conn)}
    classes = [row[0] for row in conn.execute("SELECT DISTINCT class_name FROM tt_timetable WHERE version_id=? ORDER BY class_name", (version_id,)) if row[0] in included]
    if not classes:
        raise ValueError("This version has no timetable slots.")
    if Workbook is None:
        path = _path("timetable", "xlsx")
        lines = []
        for class_name in classes:
            lines.append(f"Class Timetable — {class_name}")
            lines.extend(",".join(str(cell).replace("\n", " ") for cell in row) for row in _grid_data(conn, version_id, class_name)[0])
        Path(path).write_text("\n".join(lines), encoding="utf-8")
        return path
    workbook = Workbook()
    workbook.remove(workbook.active)
    for class_name in classes:
        title = re.sub(r"[\\/*?:\[\]]", "_", class_name)[:31] or "Class"
        sheet = workbook.create_sheet(title)
        data, _days = _grid_data(conn, version_id, class_name)
        sheet.append([SCHOOL_NAME])
        sheet.append([f"Class Timetable — {class_name}"])
        sheet.append([f"Version: {version['label']} | Academic year: {version['academic_year']}"])
        sheet.append([])
        for row in data:
            sheet.append(row)
        for cell in sheet[5]:
            cell.fill = PatternFill("solid", fgColor="5B3FC0")
            cell.font = Font(color="FFFFFF", bold=True)
        for row in sheet.iter_rows(min_row=5):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions["A"].width = 18
        for column in range(2, len(data[0]) + 1):
            sheet.column_dimensions[chr(64 + column)].width = 22
        sheet.freeze_panes = "B6"
    path = _path("timetable", "xlsx")
    workbook.save(path)
    return path
