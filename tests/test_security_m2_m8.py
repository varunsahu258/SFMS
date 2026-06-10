"""Regression tests for M-2 through M-8 security controls."""

from __future__ import annotations

import base64
import json
import os
import sqlite3
import types
from pathlib import Path

import bcrypt
import pytest
from openpyxl import load_workbook

import auth
import database
import integrity
import oauth_credentials
import ui_settings
from excel_exporter import export_to_excel
from security_utils import (
    BOOTSTRAP_PASSWORD_POLICY_MESSAGE,
    GENERIC_LOGIN_FAILURE_MESSAGE,
    display_aadhaar,
    mask_aadhaar,
    sanitize_excel_cell,
    validate_bootstrap_password,
)


def _auth_db(path: Path, password: str = "CorrectPass1!") -> None:
    with sqlite3.connect(path) as conn:
        conn.executescript(
            """
            CREATE TABLE users(
                id INTEGER PRIMARY KEY, username TEXT UNIQUE, password_hash TEXT,
                role TEXT, is_active INTEGER DEFAULT 1, failed_attempts INTEGER DEFAULT 0,
                locked_at TEXT, last_login TEXT
            );
            CREATE TABLE settings(key TEXT PRIMARY KEY,value TEXT);
            CREATE TABLE audit_log(
                id INTEGER PRIMARY KEY,timestamp TEXT,user_id INTEGER,action TEXT,table_name TEXT,
                record_id TEXT,old_value TEXT,new_value TEXT,tamper_attempt INTEGER DEFAULT 0
            );
            """
        )
        conn.execute(
            "INSERT INTO users(username,password_hash,role,is_active,failed_attempts) VALUES(?,?,?,?,0)",
            ("owner", bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode(), "ADMIN", 1),
        )


def _failure_reasons(path: Path) -> list[str]:
    with sqlite3.connect(path) as conn:
        rows = conn.execute("SELECT new_value FROM audit_log WHERE action='LOGIN_FAIL' ORDER BY id").fetchall()
    return [json.loads(row[0])["failure_reason"] for row in rows]


def test_login_failures_are_generic_and_audit_distinct_reasons(tmp_path, monkeypatch):
    db = tmp_path / "auth.db"
    _auth_db(db)
    monkeypatch.setattr(auth, "DB_PATH", db)

    ok, message = auth.login("missing", "whatever", source_ip="127.0.0.1")
    assert not ok
    assert message == GENERIC_LOGIN_FAILURE_MESSAGE

    ok, message = auth.login("owner", "wrong")
    assert not ok
    assert message == GENERIC_LOGIN_FAILURE_MESSAGE

    for _ in range(auth.MAX_FAILED_ATTEMPTS - 1):
        auth.login("owner", "wrong")
    ok, message = auth.login("owner", "wrong")
    assert not ok
    assert message == GENERIC_LOGIN_FAILURE_MESSAGE

    reasons = _failure_reasons(db)
    assert "USER_NOT_FOUND" in reasons
    assert "BAD_PASSWORD" in reasons
    assert "LOCKED" in reasons
    with sqlite3.connect(db) as conn:
        payload = json.loads(conn.execute("SELECT new_value FROM audit_log WHERE new_value LIKE '%USER_NOT_FOUND%' LIMIT 1").fetchone()[0])
    assert payload["severity"] == "WARNING"
    assert payload["username_attempted"] == "missing"
    assert payload["source_ip"] == "127.0.0.1"
    assert payload["timestamp"]


def test_bootstrap_rejects_weak_password_and_env_bootstrap_clears_secret(monkeypatch):
    conn = sqlite3.connect(":memory:")
    database._create_tables(conn)
    assert database.bootstrap_required(conn) is True
    ok, message = validate_bootstrap_password("weak")
    assert not ok
    assert message == BOOTSTRAP_PASSWORD_POLICY_MESSAGE

    monkeypatch.setenv("SFMS_BOOTSTRAP_PASSWORD", "StrongPass1!")
    user_id = database.bootstrap_from_environment(conn)
    assert user_id is not None
    assert "SFMS_BOOTSTRAP_PASSWORD" not in os.environ
    assert database.bootstrap_required(conn) is False
    assert database.bootstrap_from_environment(conn) is None


def test_create_initial_admin_policy_and_second_run_blocked():
    conn = sqlite3.connect(":memory:")
    database._create_tables(conn)
    with pytest.raises(ValueError):
        database.create_initial_admin(conn, "admin", "weak")
    admin_id = database.create_initial_admin(conn, "admin", "StrongPass1!")
    assert admin_id == 1
    with pytest.raises(ValueError):
        database.create_initial_admin(conn, "admin2", "StrongPass1!")


def _machine_conn(password="AdminPass1!"):
    conn = sqlite3.connect(":memory:")
    conn.executescript(
        """
        CREATE TABLE settings(key TEXT PRIMARY KEY,value TEXT);
        CREATE TABLE users(id INTEGER PRIMARY KEY,username TEXT,password_hash TEXT,role TEXT,is_active INTEGER);
        CREATE TABLE audit_log(id INTEGER PRIMARY KEY,timestamp TEXT,user_id INTEGER,action TEXT,table_name TEXT,record_id TEXT,old_value TEXT,new_value TEXT,tamper_attempt INTEGER);
        """
    )
    conn.execute(
        "INSERT INTO users VALUES(1,'admin',?,'ADMIN',1)",
        (bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode(),),
    )
    return conn


def test_machine_fingerprint_requires_admin_authorization(monkeypatch):
    monkeypatch.setenv("SFMS_INTEGRITY_KEY", base64.urlsafe_b64encode(b"x" * 32).decode())
    monkeypatch.setattr(integrity.socket, "gethostname", lambda: "Host One")
    monkeypatch.setattr(integrity.uuid, "getnode", lambda: 123)
    monkeypatch.setattr(integrity.platform, "machine", lambda: "x86_64")
    monkeypatch.setattr(integrity.platform, "processor", lambda: "cpu")
    conn = _machine_conn()

    assert integrity.record_machine_fingerprint(conn) is True
    stored = conn.execute("SELECT value FROM settings WHERE key='machine_id'").fetchone()[0]
    assert integrity.record_machine_fingerprint(conn) is True

    monkeypatch.setattr(integrity.socket, "gethostname", lambda: "Host Two")
    with pytest.raises(integrity.MachineAuthorizationRequired):
        integrity.record_machine_fingerprint(conn)
    assert conn.execute("SELECT value FROM settings WHERE key='machine_id'").fetchone()[0] == stored
    assert integrity.machine_authorization_required(conn) is True

    assert integrity.authorize_new_machine(conn, "admin", "AdminPass1!") is True
    assert conn.execute("SELECT COUNT(*) FROM audit_log WHERE action='MACHINE_AUTHORIZED'").fetchone()[0] == 1
    assert integrity.machine_authorization_required(conn) is False


def test_oauth_token_uses_keyring_not_database_and_export_excludes_secret(tmp_path, monkeypatch):
    stored = {}
    fake_keyring = types.SimpleNamespace(
        set_password=lambda service, user, value: stored.__setitem__((service, user), value),
        get_password=lambda service, user: stored.get((service, user)),
    )
    monkeypatch.setattr(oauth_credentials, "_keyring", lambda: fake_keyring)
    oauth_credentials.store_oauth_token('{"refresh_token":"secret"}')
    assert oauth_credentials.load_oauth_token() == '{"refresh_token":"secret"}'

    db = tmp_path / "export.db"
    reports = tmp_path / "reports"
    _export_db(db)
    with sqlite3.connect(db) as conn:
        conn.execute("INSERT INTO settings(key,value) VALUES('oauth_token','secret-token')")
    monkeypatch.setattr(ui_settings, "DB_PATH", db)
    monkeypatch.setattr(ui_settings, "REPORTS_DIR", reports)
    path = ui_settings.export_full_database_to_excel(exported_by=1, export_password="pw")
    wb = load_workbook(path, read_only=True)
    assert "oauth_token" not in [cell.value for cell in next(wb["settings"].iter_rows(max_row=1))]
    assert "secret-token" not in Path(path).read_bytes().decode("latin1", errors="ignore")


def test_excel_sanitization_and_export_cells(tmp_path, monkeypatch):
    assert sanitize_excel_cell("=SUM(A1:A10)") == "'=SUM(A1:A10)"
    assert sanitize_excel_cell("+cmd|' /C calc'!A0") == "'+cmd|' /C calc'!A0"
    assert sanitize_excel_cell("Ravi Kumar") == "Ravi Kumar"
    assert sanitize_excel_cell("Line\nBreak\x00") == "Line Break"
    monkeypatch.setattr("excel_exporter.REPORTS_DIR", tmp_path)
    path = export_to_excel([{"student_name": "=SUM(A1:A10)"}], ["student_name"], "T", "out.xlsx", {})
    wb = load_workbook(path)
    cell = wb["Data"]["A2"]
    assert cell.value == "'=SUM(A1:A10)"
    assert cell.data_type == "s"


def _export_db(path: Path) -> None:
    with sqlite3.connect(path) as conn:
        conn.executescript(
            """
            CREATE TABLE users(id INTEGER PRIMARY KEY,username TEXT,password_hash TEXT,failed_attempts INTEGER);
            CREATE TABLE students(id INTEGER PRIMARY KEY,name TEXT,aadhaar TEXT,phone TEXT);
            CREATE TABLE settings(key TEXT PRIMARY KEY,value TEXT);
            CREATE TABLE audit_log(id INTEGER PRIMARY KEY,timestamp TEXT,user_id INTEGER,action TEXT,table_name TEXT,record_id TEXT,old_value TEXT,new_value TEXT,tamper_attempt INTEGER);
            INSERT INTO users VALUES(1,'admin','hash',3);
            INSERT INTO students VALUES(1,'Ravi','123412341234','9999999999');
            INSERT INTO settings VALUES('school_name','School');
            """
        )


def test_database_export_blocks_secrets_masks_aadhaar_and_audits(tmp_path, monkeypatch):
    db = tmp_path / "export.db"
    reports = tmp_path / "reports"
    _export_db(db)
    monkeypatch.setattr(ui_settings, "DB_PATH", db)
    monkeypatch.setattr(ui_settings, "REPORTS_DIR", reports)
    path = ui_settings.export_full_database_to_excel(exported_by=1, export_password="pw")
    wb = load_workbook(path, read_only=True)
    users_header = [cell.value for cell in next(wb["users"].iter_rows(max_row=1))]
    assert "password_hash" not in users_header
    student_rows = list(wb["students"].iter_rows(values_only=True))
    assert "XXXX-XXXX-1234" in student_rows[1]
    with sqlite3.connect(db) as conn:
        row = conn.execute("SELECT new_value FROM audit_log WHERE action='DATABASE_EXPORT'").fetchone()
    assert row is not None
    payload = json.loads(row[0])
    assert payload["sha256_of_file"]


def test_aadhaar_display_masks_for_ui_contexts():
    assert display_aadhaar("123412341234", "ADMIN", "identity_verification") == "123412341234"
    assert display_aadhaar("123412341234", "ACCOUNTANT", "student_search") == "XXXX-XXXX-1234"
    assert display_aadhaar("123412341234", "ADMIN", "bulk_import_preview") == "XXXX-XXXX-1234"
    assert display_aadhaar("123", "ACCOUNTANT", "student_search") == "XXXX-XXXX-XXXX"
    assert mask_aadhaar("123412341234") == "XXXX-XXXX-1234"
